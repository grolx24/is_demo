from datetime import datetime as dt

from django.conf import settings
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def date_parse(products):
    return {
        product["ID"]: dt.strptime(product["DATE_CREATE"], "%Y-%m-%dT%H:%M:%S%z").strftime("%d.%m.%Y %H:%M:%S")
        for product in products
    }


def user_names(users):
    return {user["ID"]: f'{user["NAME"]} {user["LAST_NAME"]}' for user in users}


def save_file(products, users, file_path):
    # Создание объекта рабочей книги excel
    workbook = Workbook()
    sheet = workbook.active
    users_name = user_names(users)
    date = date_parse(products)

    # Запись данных в ячейки
    headers = [
        ["ID", "Название", "Код товара", "Дата создания", "Кем изменено", "Кем создано", "ID каталога",
         "Описание товара", "Цена", "Валюта"],
    ]
    data = headers

    for product in products:
        data.append([product['ID'], product['NAME'], product['CODE'], date[product['ID']],
                     users_name[product['MODIFIED_BY']], users_name[product['CREATED_BY']], product['CATALOG_ID'],
                     product['DESCRIPTION'], product['PRICE'], product['CURRENCY_ID']])

    for row in data:
        sheet.append(row)

    # Определение диапазона данных
    num_rows = len(data)
    num_cols = len(data[0])
    start_cell = "A1"
    end_cell = get_column_letter(num_cols) + str(num_rows + 1)

    # Создание объекта таблицы
    table = Table(displayName="Table1", ref=f"{start_cell}:{end_cell}")

    # Определение стиля таблицы
    table_style = TableStyleInfo(name="TableStyleMedium6", showFirstColumn=False,
                                 showLastColumn=False, showRowStripes=True,
                                 showColumnStripes=False)

    table.tableStyleInfo = table_style

    # Добавление таблицы на лист
    sheet.add_table(table)

    # устанавливает ширину столбцов равную максимальной ширине ячейки в столбце
    for column_cells in sheet.columns:
        try:
            max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
            column_letter = column_cells[0].column_letter
            sheet.column_dimensions[column_letter].width = max_length + 2
        except Exception as e:
            print(f"Error column_dimensions: {e}")

    # обновляет внутренние размеры листа
    sheet.calculate_dimension()

    # Сохранение файла
    try:
        workbook.save(file_path)
    except IOError as e:
        print(e)
